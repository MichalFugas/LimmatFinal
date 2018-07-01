import openpyxl,os, calendar, csv



def fileCreator(chauffeurName,monat_Nr):

    #
    global filepath_withoutEnding,filepath,dir_with_month
    monat = calendar.month_name[monat_Nr]
    filepath_withoutChauffeur = "/Users/michalfugas/Documents/Limmat/"



    filepath_withoutEnding = os.path.join("/Users/michalfugas/Documents/Limmat/Chauffeur/"+chauffeurName.replace(" ","")+"/")
    dir_with_month = os.path.join(filepath_withoutEnding+ "/" + monat+"/")


    try:
        os.stat(filepath_withoutChauffeur+"/Chauffeur/")
    except:
        os.mkdir(filepath_withoutChauffeur+"/Chauffeur/")

    try:
        os.stat(filepath_withoutEnding)
    except:
        os.mkdir(filepath_withoutEnding)



    try:
        os.stat(dir_with_month)
    except:
        os.mkdir(dir_with_month)

    filepath = os.path.join(dir_with_month + "/" + chauffeurName.replace(" ", "") + "_" + monat + ".xlsx")


    global wb
    wb = openpyxl.Workbook()
    global ws
    ws = wb.active

def exportHeader():


    header = ['Subject','Start Date','Start Time','End Date','End Time','All Day Event','Description','Location','Private']
    for i in range (len(header)):
        ws.cell(row = 1, column = i+1).value = header[i]



def valuesInserter(Subjec,StartDate,StartTime,EndDate,EndTime,AllDayEvent,Description,Location,Private,licznik):#,LK,Ao,Eo):

    wartosci = [Subjec,StartDate,StartTime,EndDate,EndTime,AllDayEvent,Description,Location,Private]

    for i in range (len(wartosci)):
        ws.cell(row = licznik, column = i+1).value = wartosci[i]

    wb.save(filepath)


def xlsxTocsv(chauffeurName,monat_Nr):
    monat = calendar.month_name[monat_Nr]
    filepath_csv = os.path.join(filepath_withoutEnding.replace(" ","")+"/"+monat+"/"+chauffeurName.replace(" ","")+"_"+monat+".csv")
    wb_csv = openpyxl.load_workbook(filename=filepath)
    print(wb_csv.worksheets)
    ws_csv = wb_csv.worksheets[0]
    print(ws_csv)
    with open(filepath_csv, 'w') as f:
        c = csv.writer(f)

        for row in ws_csv.rows:
            print([cell.value for cell in row])
            c.writerow([cell.value for cell in row])


def rchop(thestring, ending):
  if thestring.endswith(ending):
    return thestring[:-len(ending)]
  return thestring
