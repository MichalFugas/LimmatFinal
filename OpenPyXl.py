import openpyxl, os, datetime, valueOperations, re

#   Excel input file öffnen
os.chdir('/Users/michalfugas/Documents/Limmat/')
wb = openpyxl.load_workbook('/Users/michalfugas/Documents/Limmat/LimmatKalendarz/Limmat_kalendarz.xlsx')

alleSchichtliste = wb.sheetnames
ws = wb.worksheets
dws = wb.worksheets[0]
lokalizacja = 'Dietikon'
falsz = 'FALSE'
name1 = 'Limmat_1'
name2 = 'Limmat_2'
name3 = 'Limmat_3'
z = [0,0,0,0,0]

def menu():
    print ("Bitte option wählen\n"+
        "1 ->   CSV Kalender für den Chauffeur bilden\n"+
        "2 ->   CSV Kalender für alle bilden\n"+
        "3 ->   Ausdrucke wer hat an dem Tag feri\n")
    option = int(input("Alsoooo...\n"))
    if option == 1:
        option_eins()
    elif option ==2:
        option_zwei()
    elif option ==3:
        option_drei()

#### OPTION EINS

def option_eins():

    #   Alle Monate ausdrucken
    licznik = 1
    for einsazNumer in range (1, len(alleSchichtliste)):
        text = alleSchichtliste[einsazNumer]
        print (einsazNumer, text)
    print('')

    #   Monat auswahl
    monat_Nr = int(input('Bitte Monat Numer eingeben: '))
    ws = wb.worksheets[monat_Nr]
    print(alleSchichtliste[monat_Nr] +'\n')


    for einsazNumer in range (3, ws.max_row + 1):
        print(einsazNumer, ws.cell(row=einsazNumer, column=1).value)

    #   Chauffeur auswahl
    chauffeurNu = int(input('\nBitte Chauffeur wählen: '))
    chauffeurName = ws.cell(row=chauffeurNu, column=1).value
    print(chauffeurName)
    print('')
    #   File und xlsx titelreihe bilder
    valueOperations.fileCreator(chauffeurName,monat_Nr)
    valueOperations.exportHeader()

    #   Schichten Bearbeitung
    for monatsTage in range (2,ws.max_column+1):
        if ws.cell(row=chauffeurNu, column=monatsTage).value == 'RU' or ws.cell(row=chauffeurNu, column=monatsTage).value == 'FDA' or ws.cell(row=chauffeurNu, column=monatsTage).value == 'FDA' or ws.cell(row=chauffeurNu, column=monatsTage).value == 'URLb' or ws.cell(row=chauffeurNu, column=monatsTage).value == 'ZAW' or ws.cell(row=chauffeurNu, column=monatsTage).value == 'ZA' or ws.cell(row=chauffeurNu, column=monatsTage).value == 'F' or ws.cell(row=chauffeurNu, column=monatsTage).value == 'Gok' or ws.cell(row=chauffeurNu, column=monatsTage).value == 'WBin' or ws.cell(row=chauffeurNu, column=monatsTage).value == 'Kv':
            pass
            # print''
        else:

            schicht = ws.cell(row=chauffeurNu, column=monatsTage).value
            print(ws.cell(row=chauffeurNu, column=monatsTage).data_type)
            datum = datetime.date(2018, monat_Nr, int(monatsTage - 1))

            if ws.cell(row=chauffeurNu, column=monatsTage).data_type == 's':

                schicht_ohne = valueOperations.rchop(schicht,"/2")#valueOperations.rchop(str(ws.cell(row=chauffeurNu, column=monatsTage).value), "/2")
                tdw = datum.isoweekday()
            else:
                schicht_ohne = int(valueOperations.rchop(str(ws.cell(row=chauffeurNu, column=monatsTage).value),"/2"))
                if schicht_ohne < 101 :
                    tdw = datum.isoweekday()
                elif schicht_ohne > 900:
                    tdw = datum.isoweekday()
                else:
                    tdw =  int(schicht_ohne/100)

            print(datum,schicht, tdw)


            for einsazNumer in range (2, dws.max_row):
                if  dws.cell(row=einsazNumer, column=3).value == (str(schicht_ohne) + ';' + str(tdw)):

                    schichtDwsRow = dws.cell(row=einsazNumer, column=3).row
                    print (schichtDwsRow)
        #   I Teil bearbeitung
                    if dws.cell(row=einsazNumer, column=6).value == None or "/2" in str(schicht):
                        pass

                    else:
                        for j in range (4,9,1):
                            z[j-4] = dws.cell(row = einsazNumer, column=j)

                            print ('test ',j-4,' ',z[j-4].value)
                        print (" ")
                        licznik = licznik + 1
                        zestaw = "Schicht: "+str(schicht)+"\nLinie/Kurs: "+z[0].value+"\nAnfangs Ort: "+z[1].value+"\nEnde Ort: "+z[4].value
                        valueOperations.valuesInserter(name1, datum.strftime("%d.%m.%Y"), z[2].value, datum.strftime("%d.%m.%Y"), z[3].value, falsz, zestaw, lokalizacja, falsz, licznik)
        #   II Teil bearbeitung
                    if dws.cell(row=einsazNumer, column=12).value != None or "/1" in str(schicht):
                        schichtDwsRow = dws.cell(row=einsazNumer, column=3).row
                        print(schichtDwsRow)
                        for j in range(10, 15, 1):
                            z[j - 10] = dws.cell(row=einsazNumer, column=j)

                            print('test ', j - 10, ' ', z[j - 10].value)

                        if str(z[2].value).replace(":","") > str(z[3].value).replace(":",""):
                            datum2 = datum + datetime.timedelta(days=1)
                            print ('zwiekszony')
                        else:
                            datum2 = datum
                        #print (int(str(z[2].value).replace(":","")),' ',int(str(z[3].value).replace(":","")))
                        print(" ")
                        licznik = licznik + 1
                        zestaw = "Schicht: "+str(schicht)+"\nLinie/Kurs: "+z[0].value+"\nAnfangs Ort: "+z[1].value+"\nEnde Ort: "+z[4].value
                        valueOperations.valuesInserter(name2, datum.strftime("%d.%m.%Y"), z[2].value, datum2.strftime("%d.%m.%Y"), z[3].value, falsz, zestaw, lokalizacja, falsz, licznik)

        #   III Teil bearbeitung
                    if dws.cell(row=einsazNumer, column=18).value != None:
                        schichtDwsRow = dws.cell(row=einsazNumer, column=3).row
                        print(schichtDwsRow)
                        for j in range(16, 21, 1):
                            z[j - 16] = dws.cell(row=einsazNumer, column=j)

                            print('test ', j - 16, ' ', z[j - 16].value)

                        if str(z[2].value).replace(":", "") > str(z[3].value).replace(":", ""):
                            datum2 = datum + datetime.timedelta(days=1)
                            print('zwiekszony')
                        else:
                            datum2 = datum
                        # print (int(str(z[2].value).replace(":","")),' ',int(str(z[3].value).replace(":","")))
                        print(" ")
                        licznik = licznik + 1
                        zestaw = "Schicht: "+str(schicht)+"\nLinie/Kurs: "+z[0].value+"\nAnfangs Ort: "+z[1].value+"\nEnde Ort: "+z[4].value
                        valueOperations.valuesInserter(name3, datum.strftime("%d.%m.%Y"), z[2].value, datum2.strftime("%d.%m.%Y"), z[3].value, falsz, zestaw, lokalizacja, falsz, licznik)

    #   CSV bilden
    valueOperations.xlsxTocsv(chauffeurName,monat_Nr)




###### OPTION ZWEI




def option_zwei():

    #   Alle Monate ausdrucken

    for einsazNumer in range (1, len(alleSchichtliste)):
        text = alleSchichtliste[einsazNumer]
        print (einsazNumer, text)
    print('')

    #   Monat auswahl
    monat_Nr = int(input('Bitte Monat Numer eingeben: '))
    ws = wb.worksheets[monat_Nr]
    print(alleSchichtliste[monat_Nr] +'\n')


    # for einsazNumer in range (3, ws.max_row + 1):
    #     print(einsazNumer, ws.cell(row=einsazNumer, column=1).value)

    #   Chauffeur auswahl
    for chauffeurNu in range (3, ws.max_row + 1):
        licznik = 1
        chauffeurName = ws.cell(row=chauffeurNu, column=1).value
        print(chauffeurName)
        print('')
        #   File und xlsx titelreihe bilder
        valueOperations.fileCreator(chauffeurName, monat_Nr)
        valueOperations.exportHeader()

        #   Schichten Bearbeitung
        for monatsTage in range(2, ws.max_column + 1):
            if ws.cell(row=chauffeurNu, column=monatsTage).value == 'RU' or ws.cell(row=chauffeurNu,
                                                                                    column=monatsTage).value == 'FDA' or ws.cell(
                    row=chauffeurNu, column=monatsTage).value == 'FDA' or ws.cell(row=chauffeurNu,
                                                                                  column=monatsTage).value == 'URLb' or ws.cell(
                    row=chauffeurNu, column=monatsTage).value == 'ZAW' or ws.cell(row=chauffeurNu,
                                                                                  column=monatsTage).value == 'ZA' or ws.cell(
                    row=chauffeurNu, column=monatsTage).value == 'F' or ws.cell(row=chauffeurNu,
                                                                                column=monatsTage).value == 'Gok' or ws.cell(
                    row=chauffeurNu, column=monatsTage).value == 'WBin' or ws.cell(row=chauffeurNu,
                                                                                   column=monatsTage).value == 'Kv':
                pass
                # print''
            else:

                schicht = ws.cell(row=chauffeurNu, column=monatsTage).value
                print(ws.cell(row=chauffeurNu, column=monatsTage).data_type)
                datum = datetime.date(2018, monat_Nr, int(monatsTage - 1))

                if ws.cell(row=chauffeurNu, column=monatsTage).data_type == 's':

                    schicht_ohne = valueOperations.rchop(schicht,
                                                         "/2")  # valueOperations.rchop(str(ws.cell(row=chauffeurNu, column=monatsTage).value), "/2")
                    tdw = datum.isoweekday()
                else:
                    schicht_ohne = int(
                        valueOperations.rchop(str(ws.cell(row=chauffeurNu, column=monatsTage).value), "/2"))
                    if schicht_ohne < 101:
                        tdw = datum.isoweekday()
                    elif schicht_ohne > 900:
                        tdw = datum.isoweekday()
                    else:
                        tdw = int(schicht_ohne / 100)

                print(datum, schicht, tdw)

                for einsazNumer in range(2, dws.max_row):
                    if dws.cell(row=einsazNumer, column=3).value == (str(schicht_ohne) + ';' + str(tdw)):

                        schichtDwsRow = dws.cell(row=einsazNumer, column=3).row
                        print(schichtDwsRow)
                        #   I Teil bearbeitung
                        if dws.cell(row=einsazNumer, column=6).value == None or "/2" in str(schicht):
                            pass

                        else:
                            for j in range(4, 9, 1):
                                z[j - 4] = dws.cell(row=einsazNumer, column=j)

                                print('test ', j - 4, ' ', z[j - 4].value)
                            print(" ")
                            licznik = licznik + 1
                            zestaw = "Schicht: " + str(schicht) + "\nLinie/Kurs: " + z[0].value + "\nAnfangs Ort: " + z[
                                1].value + "\nEnde Ort: " + z[4].value
                            valueOperations.valuesInserter(name1, datum.strftime("%d.%m.%Y"), z[2].value,
                                                           datum.strftime("%d.%m.%Y"), z[3].value, falsz, zestaw,
                                                           lokalizacja, falsz, licznik)
                        #   II Teil bearbeitung
                        if dws.cell(row=einsazNumer, column=12).value != None or "/1" in str(schicht):
                            schichtDwsRow = dws.cell(row=einsazNumer, column=3).row
                            print(schichtDwsRow)
                            for j in range(10, 15, 1):
                                z[j - 10] = dws.cell(row=einsazNumer, column=j)

                                print('test ', j - 10, ' ', z[j - 10].value)

                            if str(z[2].value).replace(":", "") > str(z[3].value).replace(":", ""):
                                datum2 = datum + datetime.timedelta(days=1)
                                print('zwiekszony')
                            else:
                                datum2 = datum
                            # print (int(str(z[2].value).replace(":","")),' ',int(str(z[3].value).replace(":","")))
                            print(" ")
                            licznik = licznik + 1
                            zestaw = "Schicht: " + str(schicht) + "\nLinie/Kurs: " + z[0].value + "\nAnfangs Ort: " + z[
                                1].value + "\nEnde Ort: " + z[4].value
                            valueOperations.valuesInserter(name2, datum.strftime("%d.%m.%Y"), z[2].value,
                                                           datum2.strftime("%d.%m.%Y"), z[3].value, falsz, zestaw,
                                                           lokalizacja, falsz, licznik)

                        #   III Teil bearbeitung
                        if dws.cell(row=einsazNumer, column=18).value != None:
                            schichtDwsRow = dws.cell(row=einsazNumer, column=3).row
                            print(schichtDwsRow)
                            for j in range(16, 21, 1):
                                z[j - 16] = dws.cell(row=einsazNumer, column=j)

                                print('test ', j - 16, ' ', z[j - 16].value)

                            if str(z[2].value).replace(":", "") > str(z[3].value).replace(":", ""):
                                datum2 = datum + datetime.timedelta(days=1)
                                print('zwiekszony')
                            else:
                                datum2 = datum
                            # print (int(str(z[2].value).replace(":","")),' ',int(str(z[3].value).replace(":","")))
                            print(" ")
                            licznik = licznik + 1
                            zestaw = "Schicht: " + str(schicht) + "\nLinie/Kurs: " + z[0].value + "\nAnfangs Ort: " + z[
                                1].value + "\nEnde Ort: " + z[4].value
                            valueOperations.valuesInserter(name3, datum.strftime("%d.%m.%Y"), z[2].value,
                                                           datum2.strftime("%d.%m.%Y"), z[3].value, falsz, zestaw,
                                                           lokalizacja, falsz, licznik)

        #   CSV bilden
        valueOperations.xlsxTocsv(chauffeurName, monat_Nr)

def option_drei():
    datum = input('Bitte datum eingeben YYYY-MM-DD :')
    year, month, day = map(int, datum.split('-'))
    datum = datetime.date(year,month,day)
    print(datum)
    dws = wb.worksheets[month]
    print(dws)
    for dayCheck in range (1,dws.max_column):
        if day == dws.cell(row=2,column=dayCheck).value:
            for chaufferCheck in range (3,dws.max_row):
                if dws.cell(row=chaufferCheck,column=dayCheck).value == "RU" or dws.cell(row=chaufferCheck,column=dayCheck).value == "ZA":
                    print (dws.cell(row=chaufferCheck,column=1).value)




menu()