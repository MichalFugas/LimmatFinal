import openpyxl, os, datetime, valueOperations, re

os.chdir('/Users/michalfugas/Documents/')
wb = openpyxl.load_workbook('example.xlsx')

alleSchichtliste = wb.sheetnames
ws = wb.worksheets
dws = wb.worksheets[0]
lokalizacja = 'Dietikon'
falsz = 'FALSE'
name1 = 'Limmat_1'
name2 = 'Limmat_2'
name3 = 'Limmat_3'
z = [0,0,0,0,0]

licznik = 1
for i in range (1,len(alleSchichtliste)):
    text = alleSchichtliste[i]
    print (i,text)
print('')
monat_Nr = int(input('Bitte monat Numer eingeben: '))
ws = wb.worksheets[monat_Nr]
print(alleSchichtliste[monat_Nr] +'\n')


for i in range (3,ws.max_row+1):
    print(i, ws.cell(row=i, column=1).value)


chauffeurNu = int(input('\nBitte chauffeur wählen: '))
chauffeurName = ws.cell(row=chauffeurNu, column=1).value
print(chauffeurName)
print('')
valueOperations.fileCreator(chauffeurName,monat_Nr)
valueOperations.exportHeader()
for person_nummer in range (2,ws.max_column+1):
    if ws.cell(row=chauffeurNu, column=person_nummer).value == 'RU'  or ws.cell(row=chauffeurNu, column=person_nummer).value == "FDA" or ws.cell(row=chauffeurNu, column=person_nummer).value == 'URLb' or ws.cell(row=chauffeurNu, column=person_nummer).value == 'ZAW' or ws.cell(row=chauffeurNu, column=person_nummer).value == 'ZA' or ws.cell(row=chauffeurNu, column=person_nummer).value == 'F' or ws.cell(row=chauffeurNu, column=person_nummer).value == 'W' or ws.cell(row=chauffeurNu, column=person_nummer).value == 'WBin' or ws.cell(row=chauffeurNu, column=person_nummer).value == 'Kv':
        pass
        # print''
    else:

        schicht_mit_0 = ws.cell(row=chauffeurNu, column=person_nummer).value
        schicht_mit_teil = str(schicht_mit_0).lstrip("0")
        schicht = valueOperations.rchop(valueOperations.rchop(valueOperations.rchop(str(schicht_mit_0).lstrip("0"), "/1"), "/2"), "/3")
        datum = datetime.date(2018,monat_Nr,int(person_nummer-1))




        if int(valueOperations.rchop(valueOperations.rchop(valueOperations.rchop(str(ws.cell(row=chauffeurNu, column=person_nummer).value).lstrip("0"), "/1"), "/2"), "/3")) < 100 :
            tdw = datum.isoweekday()
        elif int(valueOperations.rchop(valueOperations.rchop(valueOperations.rchop(str(ws.cell(row=chauffeurNu, column=person_nummer).value).lstrip("0"), "/1"), "/2"), "/3")) > 900:
            tdw = datum.isoweekday()
        else:
            tdw = (int(schicht)/100)

        print(datum,schicht_mit_teil, int(tdw))

        for einsatz_nummer in range (2,dws.max_row):
            if dws.cell(row=einsatz_nummer, column=3).value == (str(schicht)+';'+str(tdw)):

                schichtDwsRow = dws.cell(row=einsatz_nummer, column=3).row
                print (schichtDwsRow)
#                print (ws.cell(row=chauffeurNu, column=person_nummer).value," - Numer zmiany")
#                print ('Truefalse', "/2" in str(ws.cell(row=chauffeurNu, column=einsatz_nummer).value))
                if dws.cell(row=einsatz_nummer, column=4).value == None or "/2" in str(ws.cell(row=chauffeurNu, column=person_nummer).value) or "/3" in str(ws.cell(row=chauffeurNu, column=person_nummer).value):
                    pass

                else:     # Erster Einsatz
                    for j in range (4,9,1):
                        z[j-4] = dws.cell(row = einsatz_nummer,column=j)

                        print ('test ',j-4,' ',z[j-4].value)
                    print (" ")
                    licznik = licznik + 1
                    zestaw = "Schicht: "+str(schicht)+"\nLinie/Kurs: "+z[0].value+"\nAnfangs Ort: "+z[1].value+"\nEnde Ort: "+z[4].value
                    valueOperations.valuesInserter(name1, datum.strftime("%d.%m.%Y"), z[2].value, datum.strftime("%d.%m.%Y"), z[3].value, falsz, zestaw, lokalizacja, falsz, licznik)

                if dws.cell(row=einsatz_nummer, column=10).value == None or "/1" in str(ws.cell(row=chauffeurNu, column=person_nummer).value) or "/3" in str(ws.cell(row=chauffeurNu, column=person_nummer)):
                    pass
                else:
                    schichtDwsRow = dws.cell(row=einsatz_nummer, column=3).row
                    print(schichtDwsRow)
                    for j in range(10, 15, 1):
                        z[j - 10] = dws.cell(row=einsatz_nummer, column=j)

                        print('test ', j - 10, ' ', z[j - 10].value)

                    if str(z[2].value).replace(":","") > str(z[3].value).replace(":",""):
                        datum2 = datum + datetime.timedelta(days=1)
                        print ('zwiekszony')
                    else:
                        datum2 = datum
                    #print (int(str(z[2].value).replace(":","")),' ',int(str(z[3].value).replace(":","")))
                    print(" ")
                    licznik = licznik + 1
                    zestaw = "Schicht: "+str(schicht)+"\nLinie/Kurs: "+z[0].value+"\nAnfangs Ort: "+z[1].value+"\nEnde Ort: "+z[4].value
                    valueOperations.valuesInserter(name2, datum.strftime("%d.%m.%Y"), z[2].value, datum2.strftime("%d.%m.%Y"), z[3].value, falsz, zestaw, lokalizacja, falsz, licznik)

                if dws.cell(row=einsatz_nummer, column=17).value != None: #or "/1" in str(ws.cell(row=chauffeurNu, column=person_nummer).value) or "/2" in str(ws.cell(row=chauffeurNu, column=person_nummer).value):
                    schichtDwsRow = dws.cell(row=einsatz_nummer, column=3).row
                    print(schichtDwsRow)
                    for j in range(16, 21, 1):
                        z[j - 16] = dws.cell(row=einsatz_nummer, column=j)

                        print('test ', j - 16, ' ', z[j - 16].value)

                    if str(z[2].value).replace(":", "") > str(z[3].value).replace(":", ""):
                        datum2 = datum + datetime.timedelta(days=1)
                        print('zwiekszony')
                    else:
                        datum2 = datum
                    # print (int(str(z[2].value).replace(":","")),' ',int(str(z[3].value).replace(":","")))
                    print(" ")
                    licznik = licznik + 1
                    zestaw = "Schicht: "+str(schicht)+"\nLinie/Kurs: "+z[0].value+"\nAnfangs Ort: "+z[1].value+"\nEnde Ort: "+z[4].value
                    valueOperations.valuesInserter(name3, datum.strftime("%d.%m.%Y"), z[2].value, datum2.strftime("%d.%m.%Y"), z[3].value, falsz, zestaw, lokalizacja, falsz, licznik)

#valueOperations.csv(chauffeurName)
